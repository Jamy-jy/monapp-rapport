import pandas as pd
from django.http import HttpResponse
from django.utils.timezone import make_aware
from django.utils.dateparse import parse_datetime
from .models import Stock_consommable
from .models import Imprimante
from .models import Vignettes
from mouvement_vol.models import Vol, Mouvement_vol, CategorieSejour

def export_stock_consommable(request):
    date_debut = request.GET.get("dateDebut")
    date_fin = request.GET.get("dateFin")

    date_debut = parse_datetime(date_debut)
    date_fin = parse_datetime(date_fin)

    if date_debut and date_debut.tzinfo is None:
        date_debut = make_aware(date_debut)

    if date_fin and date_fin.tzinfo is None:
        date_fin = make_aware(date_fin)

    # ========================
    # 1. STOCK CONSOMMABLE
    # ========================
    stock_query = (
        Stock_consommable.objects.select_related("consommable")
        .filter(date_mouvement__range=(date_debut, date_fin))
        .values(
            "consommable__nom_consommable",
            "qte_entree",
            "qte_sortie",
            "qte_restant",
            "date_mouvement"
        )
        .order_by("date_mouvement")
    )
    print("DEBUG DATE:", date_debut, date_fin)
    print("COUNT:", stock_query.count())

    #transformation en DataFrame

    df_stock = pd.DataFrame(list(stock_query))

    if not df_stock.empty:
        df_stock["date_mouvement"] = pd.to_datetime(df_stock["date_mouvement"]).dt.tz_localize(None)

        #rename colonnes pour Excel
        df_stock.rename(columns={
            "consommable__nom_consommable": "consommable",
            "qte_entree": "entrée",
            "qte_sortie": "sortie",
            "qte_restant": "reste",
            "date_mouvement": "date_mouvement",
        }, inplace=True)
    else:
        df_stock = pd.DataFrame(columns=[
        "consommable", "entrée", "sortie", "reste", "date_mouvement"
        ])

    # ========================
    # 2. IMPRIMANTE
    # ========================
    imprimantes = (
        Imprimante.objects
        .select_related("boxOp")
        .filter(date_prise__range=(date_debut, date_fin))
    )

    box_ops = []
    copies = []
    dates = []

    for imp in imprimantes:
        box_ops.append(imp.boxOp.numero_boxOp if imp.boxOp else "")
        copies.append(imp.nb_copie)
         # conversion timezone → Excel compatible
        date_naive = imp.date_prise.replace(tzinfo=None) if imp.date_prise else ""
        dates.append(date_naive.strftime("%Y-%m-%d"))
    
    data_imprimante = [
        ["Box Op"] + box_ops,
        ["Nombre copie"] + copies,
        ["Date de prise"] + dates,
    ]
    df_imprimante = pd.DataFrame(data_imprimante)

    # ========================
    # 3.Vignette
    # ========================
    vignettes = (
        Vignettes.objects
        .select_related("num_bobine__box_paf")
        .filter(date_modif__range=(date_debut, date_fin))
    )

    data_vignettes = []

    for v in vignettes:
        bobine = v.num_bobine
        numero_bobine = f"Bobine {bobine.numero_bobine}" if bobine else ""
        box_paf = (
            bobine.box_paf.numero_boxPaf if bobine and bobine.box_paf else ""
        )
        num_vignette = f"{bobine.debut_serie} à {bobine.fin_serie}" if bobine else ""
        nb_conso = (v.num_FinVignette - v.num_debutVignette) - v.nb_abime

        data_vignettes.append({
            "Numero Bobine": numero_bobine,
            "Box Paf": box_paf,
            "Numéro Vignette": num_vignette,
            "Numéro debut de Série": v.num_debutVignette,
            "Numéro fin de Série": v.num_FinVignette,
            "Nombre consommé": nb_conso,
            "Nombre abimé": v.nb_abime,
            "cause": v.cause,
        })
    
    df_vignettes = pd.DataFrame(data_vignettes)

    if df_vignettes.empty:
        df_vignettes = pd.DataFrame([{
            "Numero Bobine": "Aucune donnée",
            "Box Paf": "",
            "Numéro Vignette": "",
            "Numéro debut de Série": "",
            "Numéro fin de Série": "",
            "nombre consommé": "",
            "Nombre abimé": "",
            "cause": ""
        }])

    #réponse Excel
    # ========================
    # 4. ÉCRITURE EXCEL
    # ========================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="export_stock.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        sheet_name = "sheet1"
        # écrire stock
        df_stock.to_excel(
            writer, 
            index=False, 
            startrow=1, 
            sheet_name=sheet_name
            )

        worksheet = writer.sheets[sheet_name]

        worksheet.cell(row=1, column=1, value="STOCK")

        # position imprimante
        start_row_imprimante = len(df_stock) + 5

        worksheet.cell(row=start_row_imprimante, column=1, value="IMPRIMANTE")
        # écrire bloc imprimante
        if not df_imprimante.empty:
            df_imprimante.to_excel(
                writer, 
                index=False, 
                header=False, 
                startrow=start_row_imprimante,
                sheet_name=sheet_name
                )

        # VIGNETTES
        start_row_vignettes = start_row_imprimante + len(df_imprimante) + 4

        worksheet.cell(row=start_row_vignettes, column=1, value="VIGNETTE")

        df_vignettes.to_excel(
            writer,
            index=False,
            startrow=start_row_vignettes,
            sheet_name=sheet_name
        )

    return response

def export_vol(request):
    """Export des mouvements vol par période"""

    date_debut = request.GET.get("dateDebut")
    date_fin = request.GET.get("dateFin")

    date_debut = parse_datetime(date_debut)
    date_fin = parse_datetime(date_fin)

    if date_debut and date_debut.tzinfo is None:
        date_debut = make_aware(date_debut)

    if date_fin and date_fin.tzinfo is None:
        date_fin = make_aware(date_fin)

    # Récupérer tous les vols dans la période
    vols = Vol.objects.filter(
        date_arrivee_vol__range=(date_debut, date_fin)
    ).order_by('date_arrivee_vol')

    # Récupérer les catégories triées
    categories = list(CategorieSejour.objects.all().order_by('id'))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="export_vol.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        sheet_name = "Mouvements Vol"
        workbook = writer.book
        worksheet = workbook.create_sheet(sheet_name)

        # Styles
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_font = Font(bold=True, color="FF0000")

        current_row = 1

        for vol in vols:
            #  Numéro du vol
            worksheet.cell(row=current_row, column=1, value="Numéro de Vol :")
            worksheet.cell(row=current_row, column=2, value=vol.numero_vol)
            current_row += 1

            #  En-tête tableau
            worksheet.cell(row=current_row, column=1, value="Type Visa")
            worksheet.cell(row=current_row, column=2, value="Nombre")
            current_row += 1

            #  Récupérer les mouvements de ce vol
            mouvements = Mouvement_vol.objects.filter(
                num_vol=vol
            ).select_related('sejour')

            # Mapper typeVisa → nb_par_sejour
            mvt_map = {
                m.sejour.typeVisa: m.nb_par_sejour
                for m in mouvements
            }

            total = 0

            # Écrire chaque catégorie
            for cat in categories:
                qte = mvt_map.get(cat.typeVisa, 0)
                total += qte
                worksheet.cell(row=current_row, column=1, value=cat.libelle or cat.typeVisa)
                worksheet.cell(row=current_row, column=2, value=qte)
                current_row += 1

            #  Ligne total
            worksheet.cell(row=current_row, column=1, value="Total")
            worksheet.cell(row=current_row, column=2, value=total)
            current_row += 1

            # 2 lignes vides entre chaque vol
            current_row += 2

        # Supprimer la feuille par défaut si elle existe
        if "Sheet" in writer.book.sheetnames:
            del writer.book["Sheet"]

    return response
    # response = HttpResponse(
    #     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    # ) 
    # response['Content-Disposition'] = 'attachment; filename="stock_consommable.xlsx"'

    # df.to_excel(response, index=False)

    # return response